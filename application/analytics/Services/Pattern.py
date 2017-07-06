from openpyxl import Workbook
from openpyxl import load_workbook
import json

from Cell import Cell
from Models.DataTree import DataTree
from Models.Data import Data
from Functions.FunctionCatalog import FunctionCatalog
from Functions.Catalog import Catalog
from ProcedureProfiles.Interpreter import Interpreter
from Search import Search


class Pattern:

	finder = FunctionCatalog()

	def __init__(self, file: str):
		self.procedure_profile = Interpreter(file).interpret_profile()
		self.output = None
		self.coordinates = None

	def interpret(self):
		self.output = DataTree()
		# TODO : Handle parent
		parent = None
		# print("Profile: " + str(self.procedure_profile))  # TODO : FOR TEST
		count = 0
		input_dir = self.procedure_profile.pop(0)
		self.coordinates = self.procedure_profile.pop(0)
		for token in self.procedure_profile:
			# print("Token: " + token)
			is_function = Pattern.finder.get(token)
			if is_function:
				token = Catalog.catalog[token](parent)
			else:
				# TODO : What if NOT a function?
				# TODO : Token TO Sheet and Column
				temp = ""
				output = list()
				index = 0
				for character in token:
					if character == ".":
						output.insert(index, temp)
						index += 1
						temp = ""
					else:
						temp += character
				output.insert(0, temp)

				# print("Output for interpreter: " + str(output))

				file = output[1]
				# print("File: " + file)
				sheet = output[2]
				# print("Sheet: " + sheet)
				column = int(output[0])
				# print("Col: " + str(column))

				token = Data(parent, input_dir + "\\" + file, sheet, column)

			# print("New Token: " + str(token))
			self.output.add(token)
			parent = token
			# self.output.draw()
			count += 1

		return self.output

	@staticmethod
	def gen_tree(post_fix_expression):
		# print(post_fix_expression)
		data_tree = DataTree()
		root = False
		parent = None
		for node in post_fix_expression:
			token = None
			if isinstance(node, str):
				# print("String " + node)
				is_function = Pattern.finder.get(node)
				if is_function:
					token = Catalog.catalog[node](parent)
					print("Function " + str(token))
					data_tree.add(token)
				"""if not root:
					root = True
					data_tree.root = token
					print("Added " + str(token) + " as root")"""
				parent = token
				continue
			print("Adding " + str(node) + " to tree")
			data_value = Search.find(node['file'], node['sheet'], node['column'], node['object'])
			data_value = data_value[0]
			token = Cell(parent, data=data_value, file=node['file'], sheet=node['sheet'], column=node['column'])
			print("Adding token " + str(token) + " to tree")
			data_tree.add(token)
			print("Node added " + str(node))
			print("ROOT " + str(data_tree.root))

		data_tree.draw()
		print(str(data_tree.process()))

	@staticmethod
	def test_new(path="void.xls"):
		wb = Workbook()
		wb.active["A1"] = 25
		wb.save(path)

	@staticmethod
	def test_update(path="void.xls"):
		wb = load_workbook(path)
		wb.active["A1"] = 45
		wb.active["B1"] = 45
		wb.save(path)

	@staticmethod
	def test_json():
		temp = json.dumps(['foo', {'bar': ('baz', None, 1.0, 2)}])
		print(temp)
