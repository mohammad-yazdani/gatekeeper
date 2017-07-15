from Definitions import ROOT_DIR
from openpyxl import Workbook
import datetime
import os


class Export:

	def __init__(self, output: list, destination: str, coordinates: list):
		self.output_dir = ROOT_DIR + "Output\\"
		"""
		export_dir = self.output_dir + destination + datetime.datetime.now()\
			.strftime("%I:%M%p on %B %d, %Y").replace(" ", "_").replace(":", "_") + "\\"
		if not os.path.exists(export_dir):
			os.mkdir(export_dir)
		self.destination = export_dir + "result" + ".xlsx"
		print(self.destination)
		# print(str(self.output_dir))  # TODO : FOR TEST
		# print(destination)  # TODO : FOR TEST
		# print(str(output))  # TODO : FOR TEST

		self.coordinates = coordinates

		# TODO : OPEN WORKBOOK
		self.wb = Workbook()
		ws = self.wb.active

		self.title = output[0][0]
		ws.cell(row=1, column=1, value=self.title)

		self.data = output[0][1]

		# print("Output: " + str(output))
		# print("Coordinates: " + str(self.coordinates))

		row = 2
		col = 1

		for cell in self.data:
			if not cell:
				continue
			# print(str(row))
			# print(str(col))
			# print(str(cell))
			ws.cell(row=row, column=col, value=cell)
			row += 1

		print(self.destination)
		self.wb.save(filename=self.destination)
		# end """

