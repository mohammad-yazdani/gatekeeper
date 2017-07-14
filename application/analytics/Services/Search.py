import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet import worksheet

from Definitions import ROOT_DIR
from Services.ExcelInstance import ExcelInstance


class Search:

	@staticmethod
	def find_col(file, sheet, col_p):
		if not file.find(ROOT_DIR) >= 0:
			file = ROOT_DIR + "..\\files\\clientFiles\\""" + file
		# TODO : RETURN COL
		# xl = pd.ExcelFile(file)
		# df = xl.parse(sheet)
		df = pd.read_excel(file, sheet)
		for col in df:
			query = ""
			result_col = 1
			for cell in df[col]:
				query += " " + str(cell)
				if query.find(col_p) >= 0:
					result_col = str(col)
					# result_col = int(result_col.replace("Unnamed: ", "")) + 1
					return result_col

	@staticmethod
	def find_row(file, sheet, row_p):
		if not file.find(ROOT_DIR) >= 0:
			file = ROOT_DIR + "..\\files\\clientFiles\\""" + file
		# xl = pd.ExcelFile(file)
		# df = xl.parse(sheet)
		df = pd.read_excel(file, sheet)
		for col in df:
			query = ""
			for cell in df[col]:
				query += " " + str(cell)
				if query.find(row_p) >= 0:
					count = 1
					for item in df[col]:
						# print(str(count) + " " + str(item))
						if str(item).find(row_p) >= 0:
							return count - 1
						else:
							count += 1
					return -1

	@staticmethod
	def find(file, sheet, col_p, row_p):
		if not file.find(ROOT_DIR) >= 0:
			file = ROOT_DIR + "..\\files\\clientFiles\\""" + file

		try:
			open(file)
		except IOError:
			return False

		'''print("File: " + file)
		print("Sheet: " + sheet)
		print("Col: " + col_p,)
		print("Row: " + row_p, )
		print('\n')'''

		row = Search.find_row(file, sheet, row_p)
		col = Search.find_col(file, sheet, col_p)
		# print("Col: " + str(col) + " row: " + str(row))
		xl = pd.ExcelFile(file)
		df = xl.parse(sheet)
		# print(df)

		cell = [df[col][row], file]
		if cell == 'nan':
			excel_instance = ExcelInstance(file)
			cell = excel_instance.find_cell(row_p, col_p)

		return cell

	@staticmethod
	def get_row(file, sheet, row_number, ws: worksheet):
		rows = list()
		result_row = None
		for row_index in reversed(range(1, ws.max_row + 1)):
			row = list()
			for col_index in range(1, ws.max_column + 1):
				row.insert(0, ws.cell(row=row_index, column=col_index).value)
			rows.insert(0, row)
		result_row = rows[row_number - 1]
		return result_row

	@staticmethod
	def get_empty_row(file, sheet):
		wb = load_workbook(file)
		ws = wb.active

		# print(str(ws.max_row))
		# print(str(ws[1]))
		is_empty = -1
		for row in reversed(range(1, ws.max_row + 1)):
			# print("\nRow: " + str(row),)
			for item in ws[row]:
				# print(str(item.value),)
				if str(item.value).find("Total") >= 0:
					# print("TOTAL")
					is_empty = row - 1
					break
		# print(str(is_empty))
		return is_empty
