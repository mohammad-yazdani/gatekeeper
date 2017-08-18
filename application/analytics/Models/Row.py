import time
from openpyxl import load_workbook
from Services.ExcelInstance import ExcelInstance
from Definitions import ROOT_DIR
from Models.Cell import Cell
from Services.Search import Search
import os, sys, stat
import shutil


class Row:

	def __init__(self, dest_file: str, sheet: str, functions, row: int = None):
		dest_file = dest_file.replace(" ", "_")
		if not dest_file.find(ROOT_DIR) >= 0:
			dest_file = ROOT_DIR + "..\\files\\clientFiles\\""" + dest_file

		filename, file_extension = os.path.splitext(dest_file)
		# milli = int(round(time.time() * 1000))
		signature = " output"
		output_file = dest_file.replace(file_extension, signature + file_extension)

		shutil.copy(dest_file, output_file)
		self.file = output_file
		self.sheet = sheet
		# print("Row dest file: " + self.file,)
		self.cells = list()

		# TODO : Speed
		self.row = Search.get_empty_row(self.file, sheet)

		# print("Waste...", )
		self.functions = functions

	def add(self, cell: Cell):
		self.cells.insert(0, cell)

	def draw(self):
		for cell in self.cells:
			print(cell.data)

	def write(self):

		print("Writing to file: " + self.file)
		wb = load_workbook(filename=self.file, read_only=False, keep_vba=True)
		ws = wb.active

		for cell in self.cells:
			cell.write(wb, ws, dest_row=self.row)

		total_row = self.row + 1
		new_row = self.row

		for col in range(1, ws.max_column + 1):
			value = ws.cell(row=total_row, column=col).value
			value = str(value).replace(str(new_row - 1), str(new_row))
			if str(value).find("None") >= 0:
				value = ""
			ws.cell(row=total_row, column=col, value=value)

		# print("Object file: " + self.file)
		# print("Output file: " + output_file)
		# filename, file_extension = os.path.splitext(self.file)
		# self.file = self.file.replace(file_extension, "temp" + file_extension)
		wb.save(self.file)  # Write to disk

		# end """
		return self.file

	def write_functions(self):
		excel_instance = ExcelInstance(self.file)
		for method in self.functions:
			excel_instance.find_method(method)(self.row)
		excel_instance.save_and_quit()
		return self.file
