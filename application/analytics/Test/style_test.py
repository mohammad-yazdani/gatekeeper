import unittest

from openpyxl import Workbook


class MyTestCase(unittest.TestCase):
    def test_something(self):
        wb = Workbook()
        ws = wb.create_sheet(index=0, title="MyTable")
        ws['A1'].value = '3.14%'
        ws['A1'].number_format = '0.00%'
        wb.save('style_test.xlsx')


if __name__ == '__main__':
    unittest.main()
