import unittest
from openpyxl import *
import datetime


class MyTestCase(unittest.TestCase):
	def test_something(self):

		wb = load_workbook("BF_Monthly output.xlsm")

		print(wb.sheetnames)
		ws = wb.active  # Use default/active sheet
		# Or create a new named sheet and set it to active using index
		# ws = wb.create_sheet(title="User Information")
		# wb.active = 1  # Default sheet was 0, this new sheet is 1

		# cell = ws.cell(row=48, column=3, value=5000)

		total_row = 56
		new_row = 55
		for col in range(1, ws.max_column + 1):
			print(ws.cell(row=total_row, column=col).value)
		'''
		for col in range(1, ws.max_column):
			value = ws.cell(row=total_row, column=col).value
			value = str(value).replace(str(new_row - 1), str(new_row))
			ws.cell(row=total_row, column=col, value=value)

		for col in range(1, ws.max_column):
			print(ws.cell(row=total_row, column=col).value)
		# '''
		# cell.number_format = '#,##0.00€'

		# Direct cell modification
		# Add new row at bottom
		# ws.append(["1337", "NanoDano", "password1"])
		# Can use Python datetime objects

		wb.save("BF_Monthly output blah blah.xlsm")  # Write to disk

if __name__ == '__main__':
	unittest.main()
